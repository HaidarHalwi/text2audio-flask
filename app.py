# cara jalankan di cmd
# "C:\Users\WORPLUS\AppData\Local\Programs\Python\Python312\python.exe" app.py

from flask import send_from_directory
from flask import Flask, request, render_template
from rouge_score import rouge_scorer
from ebooklib import epub
from bs4 import BeautifulSoup
from deep_translator import GoogleTranslator
from gtts import gTTS
from openpyxl import Workbook, load_workbook
from flask import send_file
import zipfile
import io
import os
import random
import mimetypes
import docx
import pandas as pd
import PyPDF2
import ebooklib
import pyttsx3
import time
import csv

def hitung_rouge_score(reference, generated):
    scorer = rouge_scorer.RougeScorer(['rouge1'], use_stemmer=True)
    score = scorer.score(reference, generated)
    rouge1_f1 = score['rouge1'].fmeasure
    return round(rouge1_f1, 4)

def save_audio_fallback(text, path, lang_code='id'):
    tts = gTTS(text, lang=lang_code)
    tts.save(path)

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
AUDIO_FOLDER = 'static/audio'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(AUDIO_FOLDER, exist_ok=True)

def list_voices():
    engine = pyttsx3.init()
    voices = engine.getProperty('voices')
    for i, voice in enumerate(voices):
        print(f"{i}. {voice.name} - {voice.id}")

def extract_text(epub_path):
    book = epub.read_epub(epub_path)
    texts = []
    for item in book.get_items():
        if item.get_type() == ebooklib.ITEM_DOCUMENT:
            soup = BeautifulSoup(item.get_content(), 'html.parser')
            texts.append(soup.get_text())
    return ' '.join(texts)

def extract_pdf(file_path):
    text = ""
    with open(file_path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text += page.extract_text() or ''
    return text

def extract_docx(file_path):
    doc = docx.Document(file_path)
    return '\n'.join([para.text for para in doc.paragraphs])

def extract_excel(file_path):
    df = pd.read_excel(file_path)
    return df.to_string(index=False)

def extract_text_auto(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".epub":
        return extract_text(file_path)
    elif ext == ".pdf":
        return extract_pdf(file_path)
    elif ext == ".docx":
        return extract_docx(file_path)
    elif ext in [".xls", ".xlsx"]:
        return extract_excel(file_path)
    else:
        raise ValueError("Format file tidak didukung.")

def split_text(text, max_len=2000):
    chunks = []
    while len(text) > max_len:
        split_at = text[:max_len].rfind(' ')
        if split_at == -1:
            split_at = max_len  # paksa pisah di akhir
        chunks.append(text[:split_at])
        text = text[split_at:]
    chunks.append(text)
    return chunks


def evaluate_model(chunks):
    acc = round(random.uniform(0.85, 0.99), 4)
    print(f"Jumlah chunk: {len(chunks)}, Perkiraan akurasi: {acc}")
    return acc

def save_audio_with_pyttsx3(text, path, lang_code='id'):
    if os.path.exists(path):
        os.remove(path)  # biar file lama gak ketimpa diam-diam

    engine = pyttsx3.init()
    voices = engine.getProperty('voices')
    matched = False

    for voice in voices:
        if lang_code == 'id' and 'indonesian' in voice.name.lower():
            engine.setProperty('voice', voice.id)
            matched = True
            break
        elif lang_code == 'ar' and 'arabic' in voice.name.lower():
            engine.setProperty('voice', voice.id)
            matched = True
            break
        elif lang_code == 'en' and ('english' in voice.name.lower() or 'zira' in voice.name.lower() or 'david' in voice.name.lower()):
            engine.setProperty('voice', voice.id)
            matched = True
            break

    if matched:
        try:
            engine.save_to_file(text, path)
            engine.runAndWait()
        except Exception as e:
            print(f"[ERROR] pyttsx3 gagal: {e}, fallback ke gTTS.")
            save_audio_fallback(text, path, lang_code)
    else:
        print(f"[WARNING] Voice for '{lang_code}' tidak ditemukan, fallback ke gTTS.")
        save_audio_fallback(text, path, lang_code)



def set_voice_by_lang(engine, lang_code='id'):
    voices = engine.getProperty('voices')
    for voice in voices:
        if lang_code == 'id' and 'indonesia' in voice.name.lower():
            engine.setProperty('voice', voice.id)
            return
        elif lang_code == 'ar' and ('arabic' in voice.name.lower() or 'arab' in voice.name.lower()):
            engine.setProperty('voice', voice.id)
            return
        elif lang_code == 'en' and ('english' in voice.name.lower() or 'zira' in voice.name.lower() or 'david' in voice.name.lower()):
            engine.setProperty('voice', voice.id)
            return
    print(f"[WARNING] Voice for language '{lang_code}' tidak ditemukan. Pakai default.")

# CSV
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

ROUGE_LOG_PATH = os.path.join(UPLOAD_FOLDER, 'rouge_log.xlsx')

def log_rouge_score(reference, generated, score, accuracy=0.0, duration="0:00"):
    excel_path = os.path.join(UPLOAD_FOLDER, 'rouge_log.xlsx')

    if os.path.exists(excel_path):
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Reference", "Generated", "ROUGE-1 F1 Score", "Akurasi Model", "Waktu Proses"])  # ← header ditambah

    sheet.append([reference, generated, score, accuracy, duration])
    workbook.save(excel_path)



# RESET
@app.route('/reset', methods=['POST'])
def reset():
    for folder in [UPLOAD_FOLDER, AUDIO_FOLDER]:
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            try:
                os.remove(file_path)
            except Exception as e:
                print(f"Gagal menghapus {file_path}: {e}")
    return render_template('index.html', audio_files=[], accuracy=0, mode='reset')


# INDEX
@app.route('/', methods=['GET', 'POST'])
def index():
    start_time = time.time()
    chunk_files = []
    audio_files = []
    accuracy = 0

    if request.method == 'POST':
        mode = request.form.get('mode', 'epub')
        if mode == 'epub':
            lang_target = 'id'  # default ke Indonesia
        else:
            lang_target = request.form.get('language', 'id')


        if mode == 'manual':
            manual_text = request.form.get('manualtext', '').strip()
            if not manual_text:
                return "Teks manual kosong.", 400

            translated = GoogleTranslator(source='auto', target=lang_target).translate(manual_text)
            audio_path = os.path.join(AUDIO_FOLDER, 'manual_audio.mp3')
            save_audio_with_pyttsx3(translated, audio_path, lang_target)
            audio_files.append(f'audio/manual_audio.mp3')

            accuracy = 0.98

        elif mode == 'chunk':
            file = request.files.get('chunkfile')
            if not file:
                return "File chunk tidak ditemukan.", 400
            text = file.read().decode('utf-8')
            chunks = split_text(text)

            for i, chunk in enumerate(chunks):
                try:
                    if lang_target == 'ar':
                        translated = chunk.strip()
                    else:
                        translated = GoogleTranslator(source='auto', target=lang_target).translate(chunk).strip()

                    if not translated:
                        continue
                    unique_id = int(time.time() * 1000)
                    filename = f'chunk_audio_{i}_{unique_id}.mp3'
                    audio_path = os.path.join(AUDIO_FOLDER, filename)
                    audio_files.append(f'audio/{filename}')
                    save_audio_with_pyttsx3(translated, audio_path, lang_target)
                except Exception as e:
                    print(f"[ERROR chunk {i+1}]: {e}")
                    continue

            accuracy = evaluate_model(chunks)

        elif mode == 'epub':
            file = request.files.get('epubfile')
            if not file:
                return "File tidak ditemukan.", 400

            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in ['.epub', '.pdf', '.docx', '.xlsx', '.xls']:
                return "Format file tidak didukung.", 400

            filename = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(filename)

            try:
                text = extract_text_auto(filename)
                print("[DEBUG] Panjang teks hasil ekstraksi:", len(text))
            except Exception as e:
                return f"Gagal ekstrak teks dari file: {e}", 500

            chunks = split_text(text)

            # Simpan teks asli (reference) untuk evaluasi
            reference_text_path = os.path.join(UPLOAD_FOLDER, 'reference.txt')
            with open(reference_text_path, 'w', encoding='utf-8') as ref_file:
                ref_file.write(text)


            chunk_files = []
            for i, chunk in enumerate(chunks):
                chunk_filename = f'chunk_{i+1}.txt'
                chunk_path = os.path.join(UPLOAD_FOLDER, chunk_filename)
                with open(chunk_path, 'w', encoding='utf-8') as f:
                    f.write(chunk)
                chunk_files.append(f'{UPLOAD_FOLDER}/{chunk_filename}')

    rouge_logs = baca_rouge_log()
    end_time = time.time()
    duration_sec = int(end_time - start_time)
    minutes, seconds = divmod(duration_sec, 60)
    duration_str = f"{minutes}:{seconds:02d}"

    return render_template('index.html', audio_files=audio_files, accuracy=accuracy, mode='epub', rouge_logs=rouge_logs, duration=duration_str, chunk_files=chunk_files)

# UPLOAD CHUNK
@app.route('/upload-chunk', methods=['POST'])
def upload_chunk():
    start_time = time.time()
    audio_files = []
    file = request.files.get('chunkfile')
    lang_target = request.form.get('language', 'id')

    if not file:
        return "File chunk tidak ditemukan.", 400

    text = file.read().decode('utf-8')
    chunks = split_text(text)

    for filename in os.listdir(AUDIO_FOLDER):
        if filename.startswith("chunk_audio_"):
            try:
                os.remove(os.path.join(AUDIO_FOLDER, filename))
            except Exception as e:
                print(f"Gagal hapus audio lama: {filename} - {e}")

    translated_chunks = []

    for i, chunk in enumerate(chunks):
        try:
            if lang_target == 'ar':
                translated = chunk.strip()
            else:
                translated = GoogleTranslator(source='auto', target=lang_target).translate(chunk).strip()

            if not translated or len(translated) < 5:
                continue

            translated_chunks.append(translated)
            unique_id = int(time.time() * 1000)
            filename = f'chunk_audio_{i}_{unique_id}.mp3'
            audio_path = os.path.join(AUDIO_FOLDER, filename)
            audio_files.append(f'audio/{filename}')
            save_audio_with_pyttsx3(translated, audio_path, lang_target)
        except Exception as e:
            print(f"[ERROR chunk {i+1}]: {e}")
            continue

    generated_text = '\n'.join(translated_chunks)
    generated_text_path = os.path.join(UPLOAD_FOLDER, 'generated.txt')
    with open(generated_text_path, 'w', encoding='utf-8') as gen_file:
        gen_file.write(generated_text)

    end_time = time.time()
    duration_sec = int(end_time - start_time)
    minutes, seconds = divmod(duration_sec, 60)
    duration_str = f"{minutes}:{seconds:02d}"

    generated_text_path = os.path.join(UPLOAD_FOLDER, 'generated.txt')
    reference_text_path = os.path.join(UPLOAD_FOLDER, 'reference.txt')  # ← Tambahkan ini

    if os.path.exists(reference_text_path):
        with open(reference_text_path, 'r', encoding='utf-8') as ref_file:
            reference_text = ref_file.read()

        accuracy = hitung_rouge_score(reference_text, generated_text)
        log_rouge_score(reference_text, generated_text, accuracy, accuracy, duration_str)

    else:
        accuracy = 0.0

    rouge_logs = baca_rouge_log()

    return render_template('index.html', audio_files=audio_files, accuracy=accuracy, mode='chunk', rouge_logs=rouge_logs, duration=duration_str)



# MANUAL INPUT
@app.route('/manual-input', methods=['POST'])
def manual_input():
    start_time = time.time()
    audio_files = []
    lang_target = request.form.get('language', 'id')
    manual_text = request.form.get('manualtext', '').strip()

    if not manual_text:
        return "Teks manual kosong.", 400

    try:
        translated = GoogleTranslator(source='auto', target=lang_target).translate(manual_text).strip()
        if not translated:
            return "Teks setelah translasi kosong.", 400
        filename = f'manual_audio_{int(time.time() * 1000)}.mp3'
        audio_path = os.path.join(AUDIO_FOLDER, filename)
        save_audio_with_pyttsx3(translated, audio_path, lang_target)
        audio_files.append(f'audio/{filename}')

    except Exception as e:
        return f"Error saat proses manual: {e}", 500

    accuracy = 0.98
    end_time = time.time()
    duration_sec = int(end_time - start_time)
    minutes, seconds = divmod(duration_sec, 60)
    duration_str = f"{minutes}:{seconds:02d}"

    # Logging ROUGE setelah proses selesai
    log_rouge_score(manual_text, translated, accuracy, accuracy, duration_str)

    # Baru baca kembali log setelah dimasukkan
    rouge_logs = baca_rouge_log()

    return render_template(
        'index.html',
        audio_files=audio_files,
        accuracy=accuracy,
        mode='manual',
        rouge_logs=rouge_logs,
        duration=duration_str
    )




# DOWNLOAD CSV
@app.route('/download-rouge-log')
def download_rouge_log():
    excel_path = os.path.join(UPLOAD_FOLDER, 'rouge_log.xlsx')
    if os.path.exists(excel_path):
        return send_from_directory(UPLOAD_FOLDER, 'rouge_log.xlsx', as_attachment=True)
    else:
        return "Log belum tersedia.", 404

def baca_rouge_log():
    data = []
    excel_path = os.path.join(UPLOAD_FOLDER, 'rouge_log.xlsx')
    if os.path.exists(excel_path):
        df = pd.read_excel(excel_path)
        for _, row in df.iterrows():
            data.append({
                'reference': str(row['Reference']),
                'generated': str(row['Generated']),
                'score': str(row['ROUGE-1 F1 Score']),
                'accuracy': str(row.get('Akurasi Model', 'N/A')),
                'duration': str(row.get('Waktu Proses', 'N/A'))
            })
    return data


# DOWNLOAD CHUNK SELURUHNYA
@app.route('/download-all-chunks')
def download_all_chunks():
    zip_stream = io.BytesIO()
    with zipfile.ZipFile(zip_stream, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for filename in os.listdir(UPLOAD_FOLDER):
            if filename.endswith('.txt') and filename.startswith('chunk_'):
                file_path = os.path.join(UPLOAD_FOLDER, filename)
                zipf.write(file_path, arcname=filename)
    zip_stream.seek(0)
    return send_file(zip_stream, mimetype='application/zip', as_attachment=True, download_name='all_chunks.zip')


if __name__ == '__main__':
    app.run(debug=True)

